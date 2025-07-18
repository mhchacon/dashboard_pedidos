import openpyxl
from collections import defaultdict
from datetime import datetime

def formatar_valor(valor):
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def processar_protheus(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Encontrar os índices das colunas relevantes
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_nome = headers.index('A3_NOME')
    idx_cliente = headers.index('C5_XNOME')
    idx_valor = headers.index('C6_VALOR')
    idx_condpag = headers.index('C5_CONDPAG')
    idx_xopfat = headers.index('C5_XOPFAT')
    idx_emissao = headers.index('C5_EMISSAO')
    # Novos índices para cálculo especial
    idx_xpremio = headers.index('C6_XPREMIO') if 'C6_XPREMIO' in headers else None
    idx_qtdven = headers.index('C6_QTDVEN') if 'C6_QTDVEN' in headers else None
    idx_prcven = headers.index('C6_PRCVEN') if 'C6_PRCVEN' in headers else None
    idx_xvpremi = headers.index('C6_XVPREMI') if 'C6_XVPREMI' in headers else None
    idx_ufdest = headers.index('C5_UFDEST') if 'C5_UFDEST' in headers else None

    dados = defaultdict(lambda: defaultdict(float))
    soma_representante = defaultdict(float)
    valores_por_dia = defaultdict(float)
    valores_por_mes = defaultdict(float)
    datas = set()

    for row in ws.iter_rows(min_row=2):
        nome = row[idx_nome].value
        cliente = row[idx_cliente].value
        valor = row[idx_valor].value
        condpag = str(row[idx_condpag].value) if row[idx_condpag].value is not None else ''
        xopfat = str(row[idx_xopfat].value) if row[idx_xopfat].value is not None else ''
        data_emissao = row[idx_emissao].value
        ufdest = row[idx_ufdest].value if idx_ufdest is not None else ''
        # Filtros
        if condpag == '103':
            continue
        if xopfat not in ('1', ''):
            continue
        if not nome or not cliente or data_emissao is None:
            continue
        # Lógica para pedidos com prêmio
        valor_float = 0.0
        is_premio = False
        if idx_xpremio is not None and row[idx_xpremio].value == 'S':
            is_premio = True
            # Pega os valores, tratando nulos como zero
            qtdven = row[idx_qtdven].value if idx_qtdven is not None and row[idx_qtdven].value is not None else 0
            prcven = row[idx_prcven].value if idx_prcven is not None and row[idx_prcven].value is not None else 0
            xvpremi = row[idx_xvpremi].value if idx_xvpremi is not None and row[idx_xvpremi].value is not None else 0
            try:
                qtdven = float(str(qtdven).replace(',', '').replace('.', '.'))
            except Exception:
                qtdven = 0.0
            try:
                prcven = float(str(prcven).replace(',', '').replace('.', '.'))
            except Exception:
                prcven = 0.0
            try:
                xvpremi = float(str(xvpremi).replace(',', '').replace('.', '.'))
            except Exception:
                xvpremi = 0.0
            valor_float = (qtdven * prcven) + xvpremi
        else:
            if valor is None:
                continue
            try:
                valor_float = float(str(valor).replace(',', '').replace('.', '.'))
            except Exception:
                continue
        # Converter data para datetime
        try:
            data_str = str(data_emissao)
            if len(data_str) == 8:
                data_dt = datetime.strptime(data_str, '%Y%m%d')
            else:
                continue
        except Exception:
            continue
        dia = data_dt.strftime('%Y-%m-%d')
        mes = data_dt.strftime('%Y-%m')
        datas.add(data_dt)
        valores_por_dia[dia] += valor_float
        valores_por_mes[mes] += valor_float
        dados[nome][cliente] += valor_float
        soma_representante[nome] += valor_float

    # Montar estrutura para dashboard
    resultado = []
    for representante, clientes in dados.items():
        clientes_list = []
        for c, v in clientes.items():
            # Procurar UF do cliente
            uf_cliente = ''
            for row in ws.iter_rows(min_row=2):
                if row[idx_nome].value == representante and row[idx_cliente].value == c:
                    uf_cliente = row[idx_ufdest].value if idx_ufdest is not None else ''
                    break
            clientes_list.append({'cliente': c, 'valor': formatar_valor(v), 'UF': uf_cliente})
        # Pega o UF do primeiro cliente desse representante (assume igual para todos)
        uf_rep = ''
        for row in ws.iter_rows(min_row=2):
            if row[idx_nome].value == representante:
                uf_rep = row[idx_ufdest].value if idx_ufdest is not None else ''
                break
        resultado.append({
            'representante': representante,
            'total_representante': formatar_valor(soma_representante[representante]),
            'clientes': clientes_list,
            'C5_UFDEST': uf_rep
        })
    # Calcular último dia e mês
    if datas:
        ultimo_dia = max(datas)
        ultimo_dia_str = ultimo_dia.strftime('%Y-%m-%d')
        ultimo_mes = ultimo_dia.strftime('%Y-%m')
        valor_ultimo_dia = formatar_valor(valores_por_dia[ultimo_dia_str])
        valor_ultimo_mes = formatar_valor(valores_por_mes[ultimo_mes])
    else:
        valor_ultimo_dia = formatar_valor(0.0)
        valor_ultimo_mes = formatar_valor(0.0)
        ultimo_dia_str = ''
        ultimo_mes = ''
    return {
        'tabela': resultado,
        'ultimo_dia': ultimo_dia_str,
        'valor_ultimo_dia': valor_ultimo_dia,
        'ultimo_mes': ultimo_mes,
        'valor_ultimo_mes': valor_ultimo_mes
    }

def processar_cls(filepath):
    import openpyxl
    from collections import defaultdict
    from datetime import datetime
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_status = headers.index('Status Mov.')
    idx_favoravel = headers.index('Favorável') if 'Favorável' in headers else None
    idx_representante = headers.index('Representante')
    idx_cliente = headers.index('Nome Cliente')
    idx_valor = headers.index('Valor Líquido')
    idx_emissao = headers.index('Data Emissão')
    idx_uf = headers.index('UF') if 'UF' in headers else None

    dados = defaultdict(lambda: defaultdict(float))
    soma_representante = defaultdict(float)
    valores_por_dia = defaultdict(float)
    valores_por_mes = defaultdict(float)
    datas = set()

    for row in ws.iter_rows(min_row=2):
        status = row[idx_status].value
        favoravel = row[idx_favoravel].value if idx_favoravel is not None else 'Sim'
        representante = row[idx_representante].value
        cliente = row[idx_cliente].value
        valor = row[idx_valor].value
        data_emissao = row[idx_emissao].value
        uf = row[idx_uf].value if idx_uf is not None else ''
        if status not in ['Aprovação Comercial', 'Aprovação de Crédito', 'Em Revisão']:
            continue
        if favoravel not in ['Sim', 'S', 'Favorável', True, 1]:
            continue
        if not representante or not cliente or valor is None or not data_emissao:
            continue
        try:
            valor_float = float(str(valor).replace(',', '').replace('.', '.'))
        except Exception:
            continue
        # Converter data para datetime
        try:
            data_dt = datetime.strptime(str(data_emissao), '%d-%m-%Y')
        except Exception:
            continue
        dia = data_dt.strftime('%Y-%m-%d')
        mes = data_dt.strftime('%Y-%m')
        datas.add(data_dt)
        valores_por_dia[dia] += valor_float
        valores_por_mes[mes] += valor_float
        dados[representante][cliente] += valor_float
        soma_representante[representante] += valor_float

    resultado = []
    for representante, clientes in dados.items():
        clientes_list = []
        for c, v in clientes.items():
            # Procurar UF do cliente
            uf_cliente = ''
            for row in ws.iter_rows(min_row=2):
                if row[idx_representante].value == representante and row[idx_cliente].value == c:
                    uf_cliente = row[idx_uf].value if idx_uf is not None else ''
                    break
            clientes_list.append({'cliente': c, 'valor': formatar_valor(v), 'UF': uf_cliente})
        # Pega o UF do primeiro cliente desse representante (assume igual para todos)
        uf_rep = ''
        for row in ws.iter_rows(min_row=2):
            if row[idx_representante].value == representante:
                uf_rep = row[idx_uf].value if idx_uf is not None else ''
                break
        resultado.append({
            'representante': representante,
            'total_representante': formatar_valor(soma_representante[representante]),
            'clientes': clientes_list,
            'UF': uf_rep
        })
    # Calcular último dia e mês
    if datas:
        ultimo_dia = max(datas)
        ultimo_dia_str = ultimo_dia.strftime('%Y-%m-%d')
        ultimo_mes = ultimo_dia.strftime('%Y-%m')
        valor_ultimo_dia = formatar_valor(valores_por_dia[ultimo_dia_str])
        valor_ultimo_mes = formatar_valor(valores_por_mes[ultimo_mes])
    else:
        valor_ultimo_dia = formatar_valor(0)
        valor_ultimo_mes = formatar_valor(0)
        ultimo_dia_str = ''
        ultimo_mes = ''
    return {
        'tabela': resultado,
        'ultimo_dia': ultimo_dia_str,
        'valor_ultimo_dia': valor_ultimo_dia,
        'ultimo_mes': ultimo_mes,
        'valor_ultimo_mes': valor_ultimo_mes
    } 